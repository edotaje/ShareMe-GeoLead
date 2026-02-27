import os
import glob
import shutil
import tempfile
import threading
from collections import defaultdict
import pandas as pd
from fastapi import APIRouter, HTTPException
from fastapi.responses import FileResponse
from pydantic import BaseModel
from openpyxl import load_workbook

router = APIRouter(prefix="/api/lists", tags=["lists"])

LISTS_DIR = os.path.join(os.path.dirname(__file__), "..", "data", "lists")

# Ensure the directory exists
os.makedirs(LISTS_DIR, exist_ok=True)

# Per-file locks to prevent concurrent read-modify-write corruption
_file_locks = defaultdict(threading.Lock)

def _get_file_lock(filepath: str) -> threading.Lock:
    """Returns a lock specific to the given file path."""
    return _file_locks[os.path.normpath(filepath)]

class CreateListRequest(BaseModel):
    name: str

@router.get("/")
def get_lists():
    """Returns a list of all available Excel files in the lists directory."""
    files = glob.glob(os.path.join(LISTS_DIR, "*.xlsx"))
    # Return just the filenames without path
    file_names = [os.path.basename(f) for f in files]
    return file_names

@router.post("/")
def create_list(request: CreateListRequest):
    """Creates a new empty Excel file with the standard headers."""
    if not request.name:
        raise HTTPException(status_code=400, detail="Il nome della lista è obbligatorio")
        
    # Ensure it ends with .xlsx
    filename = request.name if request.name.endswith('.xlsx') else f"{request.name}.xlsx"
    filepath = os.path.join(LISTS_DIR, filename)
    
    if os.path.exists(filepath):
        raise HTTPException(status_code=409, detail=f"La lista '{filename}' esiste già")
        
    try:
        # Standard columns that the scraper outputs
        # We include Place_ID for deduplication
        df = pd.DataFrame(columns=[
            'Place_ID', 'Nome', 'Indirizzo', 'Telefono', 'Sito Web', 'Rating',
            'Categorie', 'Keyword Ricerca', 'Data Estrazione', 'Hide', 'Call', 'Interested', 'Note'
        ])
        df.to_excel(filepath, index=False, engine='openpyxl')
        return {"message": "Lista creata con successo", "filename": filename}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Errore durante la creazione del file: {str(e)}")

@router.get("/{filename}")
def get_list_content(filename: str):
    """Returns the content of a specific Excel list as a JSON array."""
    filepath = os.path.join(LISTS_DIR, filename)
    
    if not os.path.exists(filepath):
        raise HTTPException(status_code=404, detail="Lista non trovata")
        
    try:
        df = pd.read_excel(filepath)
        
        # Ensure new columns exist for old files
        if 'Hide' not in df.columns:
            df['Hide'] = False
        if 'Call' not in df.columns:
            df['Call'] = False
        if 'Interested' not in df.columns:
            df['Interested'] = False
        if 'Note' not in df.columns:
            df['Note'] = ''

        # Ensure NaN/NaT are converted to empty strings before sending to JSON
        # For booleans that were True/False, they might become empty strings if not careful, 
        # so let's make sure Hide/Call stay boolean
        df['Hide'] = df['Hide'].fillna(False).astype(bool)
        df['Call'] = df['Call'].fillna(False).astype(bool)
        df['Interested'] = df['Interested'].fillna(False).astype(bool)
        
        df_clean = df.fillna('')
        data = df_clean.to_dict(orient='records')
        return {"filename": filename, "data": data, "total": len(data)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Errore durante la lettura del file: {str(e)}")

@router.delete("/{filename}")
def delete_list(filename: str):
    """Deletes a specific Excel list."""
    filepath = os.path.join(LISTS_DIR, filename)
    
    if not os.path.exists(filepath):
        raise HTTPException(status_code=404, detail="Lista non trovata")
        
    try:
        os.remove(filepath)
        return {"message": f"Lista '{filename}' eliminata con successo"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Errore durante l'eliminazione del file: {str(e)}")

def _save_preserving_ricerche(filepath: str, df: pd.DataFrame):
    """Saves df to the main sheet of the Excel file, preserving the _ricerche sheet.

    Uses atomic write: writes to a temp file first, then replaces the original.
    The caller MUST hold the file lock for filepath.
    """
    # Backup _ricerche sheet if it exists
    searches_headers = []
    searches_rows = []
    try:
        searches_df = pd.read_excel(filepath, sheet_name='_ricerche')
        searches_headers = searches_df.columns.tolist()
        searches_rows = searches_df.values.tolist()
    except Exception:
        pass  # Sheet doesn't exist yet

    # Write to a temporary file first (atomic write pattern)
    dir_name = os.path.dirname(filepath)
    fd, tmp_path = tempfile.mkstemp(suffix='.xlsx', dir=dir_name)
    os.close(fd)

    try:
        # Write main sheet
        df.to_excel(tmp_path, index=False, engine='openpyxl')

        # Add _ricerche sheet if it existed
        if searches_headers:
            wb = load_workbook(tmp_path)
            ws = wb.create_sheet('_ricerche')
            ws.append(searches_headers)
            for row in searches_rows:
                ws.append(row)
            wb.save(tmp_path)

        # Atomic replace: rename temp file over original
        shutil.move(tmp_path, filepath)
    except Exception:
        # Clean up temp file on failure - original file remains intact
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
        raise


class UpdateNoteRequest(BaseModel):
    place_id: str
    note: str

@router.put("/{filename}/note")
def update_note(filename: str, request: UpdateNoteRequest):
    """Updates the Note field for a specific row in the Excel list."""
    filepath = os.path.join(LISTS_DIR, filename)

    if not os.path.exists(filepath):
        raise HTTPException(status_code=404, detail="Lista non trovata")

    try:
        with _get_file_lock(filepath):
            df = pd.read_excel(filepath)

            if 'Place_ID' not in df.columns:
                raise HTTPException(status_code=500, detail="Il file non contiene una colonna Place_ID.")

            if 'Note' not in df.columns:
                df['Note'] = ''

            mask = df['Place_ID'].astype(str) == request.place_id
            if not mask.any():
                raise HTTPException(status_code=404, detail=f"Riga con Place_ID={request.place_id} non trovata.")

            df.loc[mask, 'Note'] = request.note
            _save_preserving_ricerche(filepath, df)

        return {"message": "Nota aggiornata con successo"}
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Errore durante l'aggiornamento della nota: {str(e)}")


class UpdateRowRequest(BaseModel):
    place_id: str
    action: str # "hide", "call", or "interested"
    value: bool

@router.put("/{filename}/row")
def update_row(filename: str, request: UpdateRowRequest):
    """Updates a specific row in the Excel list based on Place_ID."""
    filepath = os.path.join(LISTS_DIR, filename)
    
    if not os.path.exists(filepath):
        raise HTTPException(status_code=404, detail="Lista non trovata")
        
    if request.action not in ["hide", "call", "interested"]:
        raise HTTPException(status_code=400, detail="Azione non valida. Usa 'hide', 'call' o 'interested'.")
        
    try:
        with _get_file_lock(filepath):
            df = pd.read_excel(filepath)

            if 'Place_ID' not in df.columns:
                raise HTTPException(status_code=500, detail="Il file non contiene una colonna Place_ID.")

            # Ensure columns exist
            if 'Hide' not in df.columns:
                df['Hide'] = False
            if 'Call' not in df.columns:
                df['Call'] = False
            if 'Interested' not in df.columns:
                df['Interested'] = False

            if request.action == "hide":
                col_name = "Hide"
            elif request.action == "call":
                col_name = "Call"
            else:
                col_name = "Interested"

            # Find the row and update
            mask = df['Place_ID'].astype(str) == request.place_id
            if not mask.any():
                raise HTTPException(status_code=404, detail=f"Riga con Place_ID={request.place_id} non trovata.")

            df.loc[mask, col_name] = request.value

            _save_preserving_ricerche(filepath, df)

        return {"message": f"Riga aggiornata con successo! {col_name}={request.value}"}
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Errore durante l'aggiornamento del file: {str(e)}")


@router.get("/{filename}/searches")
def get_searches(filename: str):
    """Returns the search history from the _ricerche sheet of the Excel list."""
    filepath = os.path.join(LISTS_DIR, filename)

    if not os.path.exists(filepath):
        raise HTTPException(status_code=404, detail="Lista non trovata")

    try:
        df = pd.read_excel(filepath, sheet_name='_ricerche')
        return df.fillna('').to_dict(orient='records')
    except Exception:
        return []  # Sheet doesn't exist yet


@router.get("/{filename}/download")
def download_list(filename: str):
    """Downloads the Excel file directly."""
    filepath = os.path.join(LISTS_DIR, filename)
    
    if not os.path.exists(filepath):
        raise HTTPException(status_code=404, detail="Lista non trovata")
        
    return FileResponse(
        path=filepath,
        filename=filename,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

